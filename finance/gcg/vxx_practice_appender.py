import pandas as pd
import numpy as np
pd.core.common.is_list_like = pd.api.types.is_list_like
import pandas_datareader as pdr
from datetime import datetime
import xlsxwriter
from any_given_week import *
from calendar import isleap

###NOTE this won't work without longer data. Goodl/yahoo are deprecated. IEX is your best shot, but they only have five years worth. &shrug 

#pegging start dates and end dates
start = datetime(2000, 10, 10)
end = datetime(2005, 10, 10)

companies = ['VXX']

#this is the code getting the data
big_daddy, lister = weekly_chance_list(companies, start, end)

#this is the chance thing, I'll change the underlting function from agw later to gradient adjut
# note: don't not adjust this - you need to add the gradient thing
chance_df = make_chancer(lister)

#big daddy is the big daddy dictionary of all the companies
writer = pd.ExcelWriter('data/vxx_tester.xlsx', engine = 'xlsxwriter')
    #this^ is pandas way of signifying you're intializing a xlxs writer
    #you have to assign it so you know the writer thing you're using

#simple for loop for saving as exel (in one sheet!)
for ticker in big_daddy.keys():
    big_daddy[ticker].to_excel(writer, sheet_name = ticker)


writer.save()

def add_years(d, years):
    new_year = d.year + years
    try:
        return d.replace(year=new_year)
    except ValueError:
        if (d.month == 2 and d.day == 29 and # leap day
            isleap(d.year) and not isleap(new_year)):
            return d.replace(year=new_year, day=28)
        raise

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

from datetime import datetime
end = datetime(2005, 10, 10)
print(end + 5)
